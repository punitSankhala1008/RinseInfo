import os
import json
from datetime import datetime
from openpyxl import Workbook

def export_json_to_excel(data: dict):
    wb = Workbook()
    ws = wb.active

    row_idx = 1
    for key, value in data.items():
        ws.cell(row=row_idx, column=1, value=key)
        if isinstance(value, (dict, list)):
            ws.cell(row=row_idx, column=2, value=json.dumps(value, indent=2))
        else:
            ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    filename = f"extracted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    downloads_dir = "downloads"
    os.makedirs(downloads_dir, exist_ok=True)  # Ensure folder exists
    file_path = os.path.join(downloads_dir, filename)
    wb.save(file_path)
    return file_path
